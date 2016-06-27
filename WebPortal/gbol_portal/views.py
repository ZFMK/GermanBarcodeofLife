from pyramid.httpexceptions import HTTPFound
from pyramid.response import Response
from pyramid.response import FileResponse
from pyramid.view import view_config
from pyramid.renderers import render
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
import pymysql
import os
from .async_email import send_mail

from .vars import config, messages


@view_config(route_name='home')
def home_view(request):
    set_language(request)
    lang = get_language(request)
    message = request.session.pop_flash()
    if len(message) < 1:
        result = render('templates/%s/home.pt' % lang, {}, request=request)
    else:
        message = message[0]
        result = render('templates/%s/home.pt' % lang, {'message': message}, request=request)
    response = Response(result)
    return response


@view_config(route_name='kontakt')
def kontakt_view(request):
    set_language(request)
    lang = get_language(request)
    conn = pymysql.connect(host=config['host'], port=config['port'],
                           user=config['user'], passwd=config['pw'], db=config['db'])
    cur = conn.cursor()
    msg = []
    if 'op' in request.POST:
        name = request.POST.get('name')
        mail = request.POST.get('mail')
        kid = request.POST.get('category')
        header = request.POST.get('header')
        text = request.POST.get('text')
        if any([name == "", mail == "", kid == "", header == "", text == ""]):
            msg.append(messages['contact'][lang])
        else:
            cur.execute("SELECT email FROM kontakt WHERE kontaktid = " + kid)
            row = cur.fetchone()
            try:
                send_mail(row[0], mail + " (" + name + ") ", header, text.replace('\t', ''))
            except Exception as e:
                msg.append('Message could not be send: %r' % e)
            else:
                msg.append(messages['contact_send'][lang])
    cur.execute("SELECT taxa, institut, name, kontaktId FROM kontakt WHERE lang='%s' ORDER BY kontaktId" % lang)
    value = []
    for row in cur.fetchall():
        value.append('<option value="{3}">{0}, {1}, {2}</option>'.format(*row))
    cur.close()
    conn.close()
    if len(msg) < 1:
        result = render('templates/%s/kontakt.pt' % lang, {'value': "".join(value)}, request=request)
    else:
        result = render('templates/%s/kontakt.pt' % lang, {'value': "".join(value),
                                                           'message': "<br />".join(msg)}, request=request)
    response = Response(result)
    return response


@view_config(route_name='links')
def links_view(request):
    set_language(request)
    lang = get_language(request)
    result = render('templates/%s/links.pt' % lang, {}, request=request)
    response = Response(result)
    return response


def previous_coll_sheet_download(request):
    templateFileName = os.path.join(config['homepath'], 'documents/download', config['collection_table']['template'])
    targetFileName = os.path.join(config['homepath'], config['collection_table']['filled'], request.params['fileName'])
    try:
        wb = load_workbook(templateFileName)
    except Exception as e:
        request.session.flash('Internal error in download_view, views.py: %r' % e)
        url = request.route_url('sammeltabelle')
        return HTTPFound(location=url)
    ws1 = wb.get_sheet_by_name('Daten')
    ws2 = wb.get_sheet_by_name('Anleitung')
    conn = pymysql.connect(host=config['host'], port=config['port'], user=config['user'], passwd=config['pw'], db=config['db'])
    cur = conn.cursor()
    cur.execute("""SELECT firstTubeId, lastTubeId, transactionkey
        FROM ShippingRequests WHERE xlsFile = "{fileName}" """.format(**request.params))
    row = cur.fetchone()
    cur.close()
    conn.close()
    i = 0
    while i <= (row[1] - row[0]):
        ws1.cell(row=i + 5, column=1).value = row[0] + i
        i += 1
    ws2.cell(row=1, column=1).value = row[2]
    wb.save(targetFileName)


@view_config(route_name='download')
def download_view(request):
    session = request.session
    set_language(request)
    lang = get_language(request)
    if "fileName" in request.params or "filename" in request.params:
        if "fileName" in request.params:
            filename = request.params['fileName']
        else:
            filename = request.params['filename']
        if "fileOption" in request.params:
            if request.params['fileOption'] == 'collectionsheet':
                filepath = os.path.join(config['homepath'], config['collection_table']['ordered'], filename)
                response = FileResponse(filepath, request=request)
                response.headers['Content-Disposition'] = ("attachment; filename=" + filename)
            elif request.params['fileOption'] == 'versandanschreiben':
                conn = pymysql.connect(host=config['host'], port=config['port'], user=config['user'],
                                       passwd=config['pw'], db=config['db'])
                cur = conn.cursor()
                cur.execute("""SELECT concat_ws(' ', salutation, title, vorname, nachname) as `name`,
                        street, concat_ws(' ', zip, city) as city, country, postaladdress,
                        s.count, transactionKey
                    From users u
                        left join ShippingRequests sr ON sr.uid = u.uid
                        Left join Shippings s on sr.id = s.ShippingRequestId
                        left join RelUserInstitution ui on sr.ContactId = ui.uid
                        left Join Institution i on i.id = ui.institutionId
                    where transactionKey = '{0}'""".format(filename))
                row = cur.fetchone()
                send_to = str(row[4]).split('\n')
                cur.close()
                if lang == 'en':
                    filename = 'cover_letter.pdf'
                else:
                    filename = 'versandanschreiben.pdf'
                coverLetterName = os.path.join(config['homepath'], 'documents/download', filename)
                c = canvas.Canvas(coverLetterName)
                c.translate(0, 700)
                c.drawImage(ImageReader(os.path.join(config['homepath'], 'static/images/logo.png')), 50, -100, 200, 200)
                c.setFont("Helvetica-Bold", 14)
                c.setFillColorRGB(0, 0, 0)
                c.drawString(20, -150, messages['letter_sender'][lang])
                c.drawString(20, -290, messages['letter_send_to'][lang])
                c.setFont("Helvetica", 14)
                c.drawString(20, -190, row[0])
                c.drawString(20, -210, row[1])
                c.drawString(20, -230, row[2])
                c.drawString(20, -250, row[3])
                i = 0
                while i < len(send_to):
                    c.drawString(20, -330 - i * 20, send_to[i])
                    i += 1
                c.setFont("Helvetica-Bold", 14)
                c.drawString(20, -370 - i * 20, messages['letter_order_no'][lang].format(row[6]))
                c.setFont("Helvetica", 14)
                c.drawString(20, -410 - i * 20, messages['letter_no_samples'][lang].format(row[5]))
                c.drawString(20, -450 - i * 20, messages['letter_body1'][lang])
                c.drawString(20, -470 - i * 20, messages['letter_body2'][lang])
                c.showPage()
                c.save()
                response = FileResponse(coverLetterName, request=request)
                response.headers['Content-Disposition'] = ("attachment; filename=%s" % filename)
                os.remove(coverLetterName)
            elif request.params['fileOption'] == 'tk_ct':  # -- called via link from viewsAdmin: sammelliste_view
                if 'role' in session and session['role'] is not None and 2 in session['role']:  # only TK allowed downl
                    response = FileResponse(
                        os.path.join(config['homepath'], config['collection_table']['filled'], filename),
                        request=request
                    )
                    response.headers['Content-Disposition'] = ("attachment; filename=" + filename)
                else:
                    url = request.route_url('sammelliste')
                    return HTTPFound(location=url)
        elif "newsUpload" in request.params:
            # used in: filemanager.js, function: selectItem and in filemanager.class.php
            filename = filename.split('?')[0].encode('utf-8')  # workaround for ckeditor always sending a timestamp with the filename
            response = FileResponse(
                os.path.join(config['homepath'], config['news']['media_directory'], filename),
                request=request
            )
            response.headers['Content-Disposition'] = ("attachment; filename=" + filename)
        elif "news" in request.params:
            response = FileResponse(
                os.path.join(config['homepath'], config['news']['media_directory'], filename),
                request=request
            )
            response.headers['Content-Disposition'] = ("attachment; filename=" + filename)
        else:
            response = FileResponse(
                os.path.join(config['homepath'], 'documents/download', filename),
                request=request
            )
            response.headers['Content-Disposition'] = ("attachment; filename=" + filename)
    return response


@view_config(route_name='impressum')
def impressum_view(request):
    set_language(request)
    lang = get_language(request)
    result = render('templates/%s/impressum.pt' % lang, {}, request=request)
    response = Response(result)
    return response


@view_config(route_name='newsletter')
def newsletter_view(request):
    set_language(request)
    lang = get_language(request)
    if 'op' in request.POST:
        conn = pymysql.connect(host=config['host'], port=config['port'], user=config['user'], passwd=config['pw'],
                               db=config['db'])
        cur = conn.cursor()
        op = request.POST.get('op')
        values = (request.POST.get('newsletters'), request.POST.get('mail'))
        print(values)
        if op in ('Abonnieren', 'Subscribe'):
            cur.execute("""INSERT INTO newsletter (type, email, uid) VALUES ("%s", "%s", null)""", values)
        else:
            cur.execute("""DELETE FROM newsletter WHERE type = "%s" AND email = "%s" """, values)
        conn.commit()
        cur.close()
        conn.close()
    result = render('templates/%s/newsletter.pt' % lang, {}, request=request)
    response = Response(result)
    return response


def set_language(request):
    session = request.session
    if 'btnGerman' in request.POST:
        session['languange'] = 'de'
    elif 'btnEnglish' in request.POST:
        session['languange'] = 'en'


def get_language(request):
    session = request.session
    if 'languange' in session:
        if session['languange'] == 'de':
            return 'de'
        elif session['languange'] == 'en':
            return 'en'
    return 'de'
