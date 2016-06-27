#!/usr/bin/python
import sys


#from ExcelObjects import *
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

import re

# for converting xls to xlsx by calling unoconv
# import subprocess

# for reading xls-file as ole object
import olefile


def checkFileExtension (filepath):

    #print (filepath)
    excelext = None
    extpat = re.compile('(\.xls[x]?$)', re.IGNORECASE)
    #extpat = re.compile('(\.xls[x]?$)|(\.ods$)', re.IGNORECASE)
    matchobj = extpat.search(filepath)
    
    if matchobj is not None:
        excelext = matchobj.group(0)
        print (excelext)
        return excelext
    else:
        raise ValueError ('Excel file has wrong file extension, must be .xls or .xlsx')
    



def readTransactionId (args):
    templateFilepath = args[0];
    
    # validate file by extension and raise error if wrong extension 
    fileext = checkFileExtension(templateFilepath)

    #get transactionId, first try if excel file could be read by openpyxl
    try:
        workbook = load_workbook(templateFilepath)
        transactionId = workbook.properties.subject

    # if not .xlsx try to read properties.subject by olefile modul  
    # xml-conversion can be done with unocnv (see below), but this requires a full openoffice installation
    except InvalidFileException:
        ole = olefile.OleFileIO(templateFilepath)
        excelmetadata = ole.get_metadata()
        transactionId = excelmetadata.subject.decode()
        print (transactionId)

    # this is the code to convert the xls to xlsx files, requires full OpenOffice or LibreOffice instance
    '''
    except InvalidFileException:
        
        convprog = 'unoconv'
        typeflag = '-f'
        typeparam = 'xlsx'
        outflag = '-o'
        outFilepath = templateFilepath + '.xlsx'
        subprocess.call([convprog, typeflag, typeparam, outflag, outFilepath, templateFilepath], timeout = 20)
        workbook = load_workbook(outFilepath)
        transactionId = workbook.properties.subject
        print (transactionId)
    '''


    return transactionId


if __name__ == "__main__":


    if len(sys.argv) != 2: 
        sys.exit ('''Get GBOL transaction id from Metadata in Collection Excel Sheet.
    Usage:        {0} "<filename>.xlxs"
    Example:
        {0} documents/download/Sammeltabelle_GBOL_2014-10-14.xlxs
        '''.format(sys.argv[0]))
        
    else:
        args = sys.argv[1:]
        transactionId = readTransactionId(args)
        exit(transactionId)

