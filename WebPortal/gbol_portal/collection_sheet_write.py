#!/usr/bin/python
import sys
import itertools

#from ExcelObjects import *
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection
from openpyxl.worksheet import SheetProtection
from openpyxl.styles import fills, colors, Color, PatternFill
from openpyxl.comments import Comment
# for tabColor? from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties

def defines():
    exceldef = {}
    exceldef['datasheet'] = 'Daten'
    exceldef['instructsheet'] = 'Anleitung'
    exceldef['firstrow'] = 5
    exceldef['cellcolor'] = Color(rgb = 'FFF9E38C') 

    return exceldef 

def setColor(cell, color):
    cell.fill = PatternFill(fill_type = 'solid', fgColor = color, bgColor = color)
    #print (cell.fill, cell.fill.fill_type)
    pass

def setDatatype(cell, coltype):
    cell.data_type = coltype['basetype']
    if coltype['num_fmt'] != None:
        cell.number_format = coltype['num_fmt']


def writeExcelFile (args, protect_headers = True, protect_ids = True, striped = True, datatypes = True, sortable = False):
    templateFilepath = args[0];
    targetFilepath = args[1];
    firstTubeNo = args[2];
    lastTubeNo = args[3];
    transactionId = args[4];
    try:
        lang = args[5]
    except IndexError:
        lang = "de"

    exceldef = defines()

    #set transactionId
    workbook = load_workbook(templateFilepath)
    workbook.properties.subject = transactionId

    #get references to sheets
    datasheet = workbook.get_sheet_by_name(exceldef['datasheet'])
    instructsheet = workbook.get_sheet_by_name(exceldef['instructsheet'])
    #helpsheet = workbook.get_sheet_by_name(exceldef['helpsheet'])

    #check datasheet size, raise error when endless rows or cols are formated
    #print (datasheet.max_column)
    if datasheet.max_column > 1023:
        raise ValueError ('excel datasheet has too much columns, check that only needed columns have been formated')
    #print (datasheet.max_row)
    if datasheet.max_row > 1023:
        raise ValueError ('excel datasheet has too much rows, check that only needed rows have been formated')        

    #enable protection to switch the feature on
    if protect_headers == True or protect_ids == True:
        sheets = workbook.get_sheet_names()
        for sheetname in sheets:
            sheet = workbook.get_sheet_by_name(sheetname)
            sheet.protection.enable()


    #shortnames for rows to insert
    firstrow = exceldef['firstrow']
    lastrow = exceldef['firstrow'] + int(lastTubeNo) - int(firstTubeNo)    

    #state of protection depends on last saved state and must be revisited for each cell
    for row in itertools.islice(datasheet.iter_rows(), 0, firstrow -1):
        for cell in row:
            if protect_headers == True and cell.comment == None: #cell protection deletes comments:
                cell.protection = Protection(locked = True, hidden = False)
            else:
                cell.protection = Protection(locked = False, hidden = False)

    #iterate through data cell and set properties
    #first read data type for column from example row
    coltype = {}
    for row in itertools.islice(datasheet.iter_rows(), firstrow -2, firstrow -1):
        for cell in row:
            #enabling arrows for column sorting in example row means to disable protection in tubenumber column and in the example fields 
            if sortable == True:
                cell.protection = Protection(locked = False, hidden = False)
            
            coltype[cell.col_idx] = {}
            coltype[cell.col_idx]['basetype'] = cell.data_type

            if coltype[cell.col_idx]['basetype'] == 'n':
                coltype[cell.col_idx]['num_fmt'] = cell.number_format
            else:
                coltype[cell.col_idx]['num_fmt'] = None
 
            #print (cell.value, coltype[cell.col_idx]['basetype'], coltype[cell.col_idx]['num_fmt'])
            
    
    #release protection for cells to be filled by collectors, apply data constraints
    colcount = datasheet.max_column
    for col in range (1, colcount + 1, 1):
        for row in range (firstrow, lastrow + 1, 1):
            cell = datasheet.cell(column = col, row = row)
            cell.protection = Protection(locked = False, hidden = False)

            if datatypes == True:
                setDatatype(cell, coltype[col])

            if col == 1: #set color to each cell of tubenumber column
                setColor(cell, exceldef['cellcolor'])
            elif striped == True and int(row) % 2 == 0: # set colors for data cells, striped 
                setColor(cell, exceldef['cellcolor'])
            
    #insert tubenumbers to column 1 in datasheet
    # do this after data_type was set  
    rownum = firstrow
    for tube in range (int(firstTubeNo), int(lastTubeNo) + 1, 1):
        cell = datasheet.cell(column = 1, row = rownum)
        cell.value = tube
        rownum += 1
        #set protection, but it is only active when sheet protection is set above
        if protect_ids == True:
            cell.protection = Protection(locked = True, hidden = False)

    #reset style and content for rows not used by tubes
    rowcount = datasheet.max_row
    for row in itertools.islice(datasheet.iter_rows(), lastrow, rowcount):
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type = None)
            cell.protection = Protection(locked = False, hidden = False)


    workbook.save(targetFilepath)
    return targetFilepath
        
        
    


if __name__ == "__main__":


    if len(sys.argv) < 5: 
        sys.exit ('''Write GBOL Collection Sheet.
    Usage:        {0} templateFilepath targetFilepath firstTubeNo lastTubeNo transactionId lang
        templateFilepath:   Path and filename of template collection sheet, e.g. documents/download/Sammeltabelle_GBOL_2014-10-14.xls
        targetFilepath:     Path and filename of desired collection sheet, e.g. documents/download/700_2015-02-18_000000.xls
        firstTubeNo:        Number of the first tube, e.g. 3500760
        lastTubeNo:         Number of the last tube, e.g. 3500854
        transactionId:      Transaction id as generated during material order, e.g. 700_2015-02-18_000000
        lang:               Current language (not used)

    Example:
        {0} documents/download/Sammeltabelle_GBOL_2014-10-14.xls documents/download/700_2015-02-18_000000.xls 3500760 3500854 700_2015-02-18_000000
    '''.format(sys.argv[0]))

    else:
        args = sys.argv[1:]
        filepath = writeExcelFile(args)

